import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict
import os
import re
import json

routes = []
months = {
    1: 'Январь',
    2: 'Февраль',
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь',
    12: 'Декабрь'
}


class ReportCreater:

    def __init__(self, report_path):
        self.report_path = report_path
        self.dict_axapta = defaultdict(lambda: defaultdict(dict))
        self.dict_organizer = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        self.period = None
        self.problems = {'БО': 'L', 'Несоблюдение трассы маршрута': 'P', 'Пропуск остановочных пунктов': 'Q',
                         'Несоблюдение времени отправления от начальных/и или конечных пунктов': 'T',
                         'Несоблюдение времени начала/окончания движения': 'R', 'ДТП': 'K', 'Т': 'G',
                         'Э': 'H', 'Проезд ОП (>20%)': 'S', 'Задержка в пути': 'J', 'Прочее': 'I', 'Трасса (>20%)': 'U'
                         }
        self.columns = ('Дата', 'Маршрут', 'План ОП', 'План наш', 'Факт ОРГП', 'Факт ВС', 'T', 'Э', 'Прочее', 'Задержки в пути',
                        'ДТП', 'БО', 'Зачтено БО', 'Не зачтено БО', 'Примечания', 'Несобл трассы', 'Пропуск ОП',
                        'Несобл нач/оконч', 'Проезд ОП (>20%)', 'Несобл времени отпр', 'Трасса (>20%)')
        self.get_routes()

    # def read_axapta_report(self):
    #     """Считывает данные отчета по тр и заполняет словарь 'dict_axapta'"""
    #
    #     print('Считываем данные из отчёта по транспортной работе...')
    #     wb = openpyxl.load_workbook(self.report_path)
    #     sheet = wb.active
    #     for row in sheet.values:
    #         if type(row[0]) == datetime:
    #             if self.period is None:
    #                 self.period = (row[0].year, row[0].month)
    #             date_, route, flights = row[0], row[2], int(row[8])
    #             self.dict_axapta[route][date_] += flights
    #     print('Данные из отчёта по транспортной работе получены!')

    def read_tw_report(self):
        """ Читает отчет по транспортной работе и заполняет словарь self.dict_axapta """

        print('Читаем "отчёт по ТР.xlsx"....')
        wb = openpyxl.load_workbook(self.report_path)
        sheet = wb['Осн']

        for row in sheet.values:
            date, route = row[0], row[4],
            if type(date) == datetime and route in routes:
                if self.period is None:
                    self.period = (date.year, date.month)

                self.dict_axapta[route][date] = {
                                                'plan_routes': row[9],
                                                'fact_routes': row[10],
                                                }

        print('Отчёт по ТР.xlsx прочитан!')

    def gen_dates(self):
        """Функция-генератор дат"""
        year, month = self.period
        for day in range(1, 32):
            try:
                cur_date = datetime(int(year), int(month), day)
                yield cur_date
            except ValueError:
                break

    def write_data(self):
        """Записывает прочитанные данные в эксель 'report.xlsx' """
        print('Составляем сводный отчёт...')
        try:
            wb = openpyxl.load_workbook('report.xlsx')
            sheet = wb.active
            self.append_to_report(sheet)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            self.create_new_report(sheet)
        wb.save('report.xlsx')
        print('Готово!')

    def create_new_report(self, sheet):
        """Создает новый отчёт 'report.xlsx'"""
        row_numb = 1
        sheet.append(self.columns)
        for route in self.dict_axapta:
            for date_ in self.gen_dates():
                row_numb += 1
                values = (date_.strftime('%d.%m.%Y'), route, self.dict_organizer[route][date_]['plan'],
                          self.dict_axapta[route][date_]['plan_routes'], self.dict_organizer[route][date_]['fact'],
                          self.dict_axapta[route][date_]['fact_routes'],
                          )
                sheet.append()
                if values[4] < values[5] and values[2] + values[4] != 0:
                    self.color_cells(sheet, row_numb)
                for problem in self.dict_organizer[route][date_]:
                    if problem in self.problems:
                        sheet[f'{self.problems[problem]}{row_numb}'] = self.dict_organizer[route][date_][problem]

    def append_to_report(self, sheet): # надо посмотреть
        """Добавляет и перезаписывает данные в старом отчёте 'report.xlsx'"""

        for row_numb in range(2, sheet.max_row + 1):
            route = sheet[f'B{row_numb}'].value
            date_str = sheet[f'A{row_numb}'].value
            date_ = datetime.strptime(date_str, '%d.%m.%Y') if type(date_str) == str else date_str
            old_values = (sheet[f'C{row_numb}'].value, sheet[f'E{row_numb}'].value,
                          sheet[f'F{row_numb}'].value)
            values = (self.dict_organizer[route][date_]['plan'], self.dict_organizer[route][date_]['fact'],
                      self.dict_axapta[route][date_]['fact_routes'], self.dict_axapta[route][date_]['plan_routes']
                      )
            sheet[f'C{row_numb}'].value = values[0]
            sheet[f'D{row_numb}'].value = values[3]
            sheet[f'E{row_numb}'].value = values[1]
            sheet[f'F{row_numb}'].value = values[2]
            if self.check_difference(values):
                self.color_cells(sheet, row_numb)
            if self.check_difference(old_values) and not self.check_difference(values):
                self.color_cells(sheet, row_numb, color=False)
            for problem in self.dict_organizer[route][date_]:
                if problem in self.problems:
                    sheet[f'{self.problems[problem]}{row_numb}'] = self.dict_organizer[route][date_][problem]

    @staticmethod
    def check_difference(values):
        return values[1] < values[2] and values[0] + values[1] != 0

    def get_organizer_data(self):
        """Получает данные из актов организатора"""
        pattern = re.compile(r'\d+[А-Яа-я]*')
        for file_name in filter(lambda f: f.endswith('xlsx'), os.listdir('org_reports')):
            route = re.search(pattern, file_name).group()
            route_code = self.get_route_code(route)
            if route_code in self.dict_organizer:
                continue
            try:
                wb = openpyxl.load_workbook(rf'org_reports\{file_name}')
                sheet = self.find_sheet(wb)
            except FileNotFoundError:
                continue
            for cell in sheet['A']:
                flight_date = cell.value
                if type(flight_date) == datetime:
                    self.add_value(route_code, sheet, cell.row, flight_date)
                    self.find_problems(route_code, sheet, cell.row, flight_date)
                if flight_date == 'ИТОГО:':
                    break
            print(f'Акт по маршруту {route_code} проверен')
        print('Все акты проверены!')

    @staticmethod
    def color_cells(sheet, row, color=True):
        """Окрашивает диапазон ячеек"""
        if color:
            filling = PatternFill(fill_type='solid', fgColor='F4A460')
        else:
            filling = PatternFill(fill_type='none')

        for cell in sheet[f'A{row}': f'F{row}'][0]:
            cell.fill = filling

    def find_sheet(self, wb):
        """Находит лист в акте ОРГП за нужный период"""

        year, month = self.period
        period = f'{months[int(month)]} {int(year) % 100}'
        for sheet in wb.sheetnames:
            if sheet.strip() == period:
                return wb[sheet]

    def find_problems(self, route_code, sheet, row, flight_date):
        """Находит нарушения, причины срывов и заносит в 'dict_organizer' """

        values = [cell.value for cell in sheet[f'M{row}': f'O{row}'][0]]
        sum_flights = sum(int(val.value) for val in sheet[f'P{row}': f'Q{row}'][0] if val.value)
        sum_flights = 1 if not sum_flights else sum_flights
        for val in values:
            if val in self.problems:
                self.dict_organizer[route_code][flight_date][val] += sum_flights
                break

    def add_value(self, route_code, sheet, row, flight_date):
        """Добавляет факт и план рейсов в 'dict_organizer'"""

        values = [self.get_int(cell.value) for cell in sheet[f'B{row}': f'F{row}'][0] if cell.column != 4]
        fact_value = int(values[1]) + int(values[3])
        plan_value = int(values[0]) + int(values[2])
        if fact_value:
            self.dict_organizer[route_code][flight_date]['fact'] = fact_value
        if plan_value:
            self.dict_organizer[route_code][flight_date]['plan'] = plan_value

    @staticmethod
    def get_int(value):
        """Возвращает целое число, если можно взять int(), иначе - 0"""
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    def get_route_code(self, route_numb: str):
        """Возвращает код маршрута, добавляя к значению 10000"""
        if not route_numb.isdigit():
            route_code = f'{10 ** 4 + int(route_numb[:-1])}{route_numb[-1]}'.upper()
            return self.check_exception(route_code)
        return f'{10 ** 4 + int(route_numb)}'

    @staticmethod
    def check_exception(route_code: str):
        if 'Э' in route_code:
            return route_code.replace('Э', '')
        return route_code

    @staticmethod
    def get_routes():
        global routes
        with open(r'routes.json', encoding='utf-8') as f:
            routes = json.load(f)


try:
    for file in os.listdir():
        if file.endswith('.xlsx') and 'tmp' in file:
            rep_name = file
            break
    rep = ReportCreater(rep_name)
    rep.read_tw_report()
    rep.get_organizer_data()
    rep.write_data()
except Exception as err:
    print(err)
    input()
